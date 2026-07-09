import csv
from html import unescape
import json
import os
import re
import ssl
from urllib.parse import parse_qs, urlencode, urljoin, urlparse
from urllib.error import URLError
from urllib.request import Request, urlopen
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

import pdfplumber


PDF_FOLDER = "ride_times"
CLASS_SCHEDULE_FOLDER = "class_schedules"
RIDERS_FILE = "riders.txt"
OUTPUT_FOLDER = "output"
SUPPORTING_OUTPUT_FOLDER = "supporting_files"
HORSE_SHOW_OFFICE_BASE_URL = "https://www.horseshowoffice.com"
FOXVILLAGE_BASE_URL = "https://www.foxvillage.com"
EQUESTRIAN_HUB_GRAPHQL_URL = (
    "https://spectatorjudginga14295f70.hana.ondemand.com/"
    "andromeda-1.0.0/api/graph"
)
EQUESTRIAN_HUB_MASTERLIST_CACHE = {}
LETTER_ONLY_CLASS_CODES = {
    "L",
    "DHGEF",
}


DAY_ORDER = {
    "Mon": 1,
    "Tue": 2,
    "Wed": 3,
    "Thu": 4,
    "Fri": 5,
    "Sat": 6,
    "Sun": 7,
}

DAY_NAMES = {
    "Monday": "Mon",
    "Tuesday": "Tue",
    "Wednesday": "Wed",
    "Thursday": "Thu",
    "Friday": "Fri",
    "Saturday": "Sat",
    "Sunday": "Sun",
}


SHOW_NAME = "Barn Schedule"


def slugify_filename(text):
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = text.strip("-")
    return text or "show"


def output_path(filename):
    show_slug = slugify_filename(SHOW_NAME)
    return os.path.join(OUTPUT_FOLDER, f"{show_slug}_{filename}")


def supporting_output_path(filename):
    show_slug = slugify_filename(SHOW_NAME)
    folder = os.path.join(OUTPUT_FOLDER, SUPPORTING_OUTPUT_FOLDER)
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, f"{show_slug}_{filename}")


arena_pattern = re.compile(r"\s(?P<arena_num>\d+):\s")
named_arena_pattern = re.compile(
    r"\s(?P<arena>(?:East|West|North|South|Main|Indoor|Outdoor|Covered|Warm[- ]?Up|"
    r"Lower|Upper|Center|Centre|Ring\s+\d+|Arena\s+\d+|[A-Z])\s+"
    r"(?:Ring|Arena|Court|Stadium|Field)\s*(?:\([^)]*\))?.*)$"
)
ride_pattern = re.compile(r"^(Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+\d")
rider_pattern = re.compile(r"^[A-Z][a-zA-Z'’\-\s]+,\s+[A-Z]")

# Finds lines in the class schedule that start with a time.
class_schedule_time_pattern = re.compile(r"^\d{1,2}:\d{2}\s+(AM|PM)\s+")

# Finds likely class codes such as H1PSG, 121, 1FFS, 2I1, OB29, PB45GS, DHPEF, US4EF.
class_code_pattern = re.compile(
    r"\b("
    r"[A-Z]*\d[A-Z0-9]*(?:-\d+)?"
    r"|PB\d+[A-Z]*"
    r"|OB\d+[A-Z]*"
    r"|DH[A-Z]+EF"
    r"|US\dEF"
    r"|L"
    r")\b"
)


def load_riders():
    if not os.path.exists(RIDERS_FILE):
        print(f"Missing rider file: {RIDERS_FILE}")
        return []

    with open(RIDERS_FILE, "r") as file:
        riders = [
            line.strip()
            for line in file
            if line.strip()
        ]

    return riders


def extract_lines_from_folder(folder):
    lines = []

    if not os.path.exists(folder):
        return lines

    for file in os.listdir(folder):
        if not file.lower().endswith(".pdf"):
            continue

        path = os.path.join(folder, file)

        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()

                if text:
                    lines.extend(text.split("\n"))

    return lines

def extract_riders_from_lines(lines):
    riders = []

    for line in lines:
        line = line.strip()

        if rider_pattern.match(line):
            if line not in riders:
                riders.append(line)

    return sorted(riders, key=lambda name: name.lower())

def clean_class_name(name):
    name = name.strip()

    # Remove schedule housekeeping words.
    name = name.replace("Continues", "").strip()

    # Remove common judge/location text that leaks into the class name.
    name = re.sub(r"\b[A-Z][a-zA-Z'’\-]+\s+[A-Z][a-zA-Z'’\-]+\s+at\s+[A-Z]\b", "", name)
    name = re.sub(r"\bat\s+[A-Z]\b", "", name)

    # Clean up repeated spaces.
    name = re.sub(r"\s+", " ", name).strip()

    return name


def group_words_into_rows(words, tolerance=3):
    rows = []

    for word in sorted(words, key=lambda w: (w["top"], w["x0"])):
        if not rows:
            rows.append([word["top"], [word]])
            continue

        previous_top = rows[-1][0]

        if abs(word["top"] - previous_top) <= tolerance:
            rows[-1][1].append(word)
        else:
            rows.append([word["top"], [word]])

    return rows


def looks_like_class_code(text):
    text = text.strip()

    if text in {"AM", "PM"}:
        return False

    if text.startswith("-"):
        return False

    if text in LETTER_ONLY_CLASS_CODES:
        return True

    # Most class codes contain at least one number:
    # H1PSG, 121, 1FFS, 2I1, OB29, PB45GS, DHPEF, US4EF, etc.
    if re.match(r"^[A-Z]*\d[A-Z0-9]*(?:-\d+)?$", text):
        return True

    return False


def parse_class_header_line(line):
    match = re.match(
        r"^Class\s+(?P<code>[A-Z0-9-]+)\s*,\s*(?P<name>.+)$",
        line.strip()
    )

    if not match:
        return "", ""

    class_code = match.group("code").strip()
    class_name = clean_class_name(match.group("name"))

    return class_code, class_name


def build_class_map():
    class_map = {}

    if not os.path.exists(CLASS_SCHEDULE_FOLDER):
        return class_map

    for file in os.listdir(CLASS_SCHEDULE_FOLDER):
        if not file.lower().endswith(".pdf"):
            continue

        path = os.path.join(CLASS_SCHEDULE_FOLDER, file)

        file_header_class_map = {}
        file_legacy_class_map = {}

        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                words = page.extract_words()
                rows = group_words_into_rows(words)

                for _, row_words in rows:
                    row_words = sorted(row_words, key=lambda w: w["x0"])

                    row_text = " ".join(w["text"] for w in row_words)

                    header_class_code, header_class_name = parse_class_header_line(row_text)
                    if header_class_code and header_class_name:
                        file_header_class_map[header_class_code] = header_class_name
                        continue

                    if "Break" in row_text or "Lunch" in row_text or "Arena Done" in row_text:
                        continue

                    # A class schedule row should begin with a time.
                    if len(row_words) < 4:
                        continue

                    if not re.match(r"^\d{1,2}:\d{2}$", row_words[0]["text"]):
                        continue

                    if row_words[1]["text"] not in {"AM", "PM"}:
                        continue

                    # In this PDF, the class code is usually the first code-like token
                    # after the time and AM/PM.
                    class_code = ""
                    class_code_index = None

                    for i, word in enumerate(row_words[2:], start=2):
                        if looks_like_class_code(word["text"]):
                            class_code = word["text"]
                            class_code_index = i
                            break

                    if not class_code:
                        continue

                    class_name_words = []

                    for word in row_words[class_code_index + 1:]:
                        text = word["text"]

                        # Judge/location text is usually far to the right.
                        # Stop before it.
                        if word["x0"] > 390:
                            break

                        if text == "Continues":
                            continue

                        class_name_words.append(text)

                    class_name = " ".join(class_name_words).strip()
                    class_name = clean_class_name(class_name)

                    if class_code and class_name:
                        existing = file_legacy_class_map.get(class_code)

                        # Prefer the shorter/base name over a messy duplicate.
                        if not existing or len(class_name) < len(existing):
                            file_legacy_class_map[class_code] = class_name

        if file_header_class_map:
            class_map.update(file_header_class_map)
        else:
            class_map.update(file_legacy_class_map)

    return class_map


def build_class_schedule_ride_lookup():
    ride_lookup = {}

    if not os.path.exists(CLASS_SCHEDULE_FOLDER):
        return ride_lookup

    for file in os.listdir(CLASS_SCHEDULE_FOLDER):
        if not file.lower().endswith(".pdf"):
            continue

        path = os.path.join(CLASS_SCHEDULE_FOLDER, file)
        current_day = ""
        current_arena = ""
        current_class_code = ""

        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""

                for raw_line in text.split("\n"):
                    line = raw_line.strip()

                    day_arena_match = re.match(
                        r"^(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday),"
                        r".*?\d{4}\s+(?P<arena>.+?(?:Ring|Arena|Court|Stadium|Field))$",
                        line
                    )

                    if day_arena_match:
                        current_day = DAY_NAMES.get(day_arena_match.group(1), "")
                        current_arena = day_arena_match.group("arena").strip()
                        continue

                    class_code, _ = parse_class_header_line(line)

                    if class_code:
                        current_class_code = class_code
                        class_lookup_key = (current_class_code, "")
                        ride_lookup.setdefault(class_lookup_key, {
                            "day": current_day,
                            "arena": current_arena,
                            "arena_number": "",
                            "arena_name": current_arena,
                        })
                        continue

                    if not current_class_code or not current_arena:
                        continue

                    if "Break" in line or "Arena Done" in line:
                        continue

                    ride_match = re.match(
                        r"^\d{1,2}:\d{2}\s+(?:AM|PM)\s+\d+\s+(?P<body>.+)$",
                        line
                    )

                    if not ride_match:
                        continue

                    body = ride_match.group("body")
                    horse = body.split(",", 1)[0].strip()

                    if not horse:
                        continue

                    lookup_key = (
                        current_class_code,
                        normalize_lookup_text(horse)
                    )
                    ride_lookup[lookup_key] = {
                        "day": current_day,
                        "arena": current_arena,
                        "arena_number": "",
                        "arena_name": current_arena,
                    }

    return ride_lookup


def enrich_rides_from_class_schedule(rides, schedule_ride_lookup):
    for ride in rides:
        lookup_key = (
            ride.get("class", ""),
            normalize_lookup_text(ride.get("horse", ""))
        )
        arena_info = schedule_ride_lookup.get(lookup_key)

        if not arena_info:
            arena_info = schedule_ride_lookup.get((ride.get("class", ""), ""))

        if not arena_info:
            continue

        ride["arena"] = arena_info["arena"]
        ride["arena_number"] = arena_info["arena_number"]
        ride["arena_name"] = arena_info["arena_name"]

    return rides

def export_class_map_csv(class_map):
    output_file = supporting_output_path("class_definitions.csv")
    
    with open(output_file, "w", newline="") as file:
        writer = csv.writer(file)

        writer.writerow([
            "Class #",
            "Class Name"
        ])

        for class_code in sorted(class_map.keys()):
            writer.writerow([
                class_code,
                class_map[class_code]
            ])

    print(f"Class map exported to: {output_file}")


def normalize_text(text):
    text = text.lower()
    text = text.replace(".", "")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_lookup_text(text):
    text = unescape(text or "")
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip().lower()


def html_cell_text(text):
    text = unescape(text or "")
    text = text.replace("\xa0", " ")
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def hso_request(url, data=None):
    encoded_data = None

    if data is not None:
        encoded_data = urlencode(data).encode()

    request = Request(
        url,
        data=encoded_data,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Content-Type": "application/x-www-form-urlencoded",
        }
    )

    try:
        with urlopen(request, timeout=30) as response:
            return response.read().decode("windows-1252", errors="replace")
    except (ssl.SSLCertVerificationError, URLError) as error:
        if isinstance(error, URLError) and not isinstance(
            error.reason,
            ssl.SSLCertVerificationError
        ):
            raise

        # HorseShowOffice currently works in browsers but may fail Python's
        # local cert chain on macOS. Retry read-only fetches with SSL relaxed.
        context = ssl._create_unverified_context()
        with urlopen(request, timeout=30, context=context) as response:
            return response.read().decode("windows-1252", errors="replace")


def url_request_text(url, encoding="utf-8", data=None):
    encoded_data = None

    if data is not None:
        encoded_data = urlencode(data).encode()

    request = Request(
        url,
        data=encoded_data,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "application/json,text/html",
        }
    )

    try:
        with urlopen(request, timeout=30) as response:
            return response.read().decode(encoding, errors="replace")
    except (ssl.SSLCertVerificationError, URLError) as error:
        if isinstance(error, URLError) and not isinstance(
            error.reason,
            ssl.SSLCertVerificationError
        ):
            raise

        context = ssl._create_unverified_context()
        with urlopen(request, timeout=30, context=context) as response:
            return response.read().decode(encoding, errors="replace")


def url_request_json(url):
    return json.loads(url_request_text(url, encoding="utf-8"))


def json_request(url, payload, headers=None):
    request_headers = {
        "User-Agent": "Mozilla/5.0",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }

    if headers:
        request_headers.update(headers)

    request = Request(
        url,
        data=json.dumps(payload).encode(),
        headers=request_headers
    )

    try:
        with urlopen(request, timeout=30) as response:
            return json.loads(response.read().decode("utf-8", errors="replace"))
    except (ssl.SSLCertVerificationError, URLError) as error:
        if isinstance(error, URLError) and not isinstance(
            error.reason,
            ssl.SSLCertVerificationError
        ):
            raise

        context = ssl._create_unverified_context()
        with urlopen(request, timeout=30, context=context) as response:
            return json.loads(response.read().decode("utf-8", errors="replace"))


def download_url_to_file(url, destination_path):
    request = Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0",
        }
    )

    destination_path = str(destination_path)

    try:
        with urlopen(request, timeout=30) as response:
            data = response.read()
    except (ssl.SSLCertVerificationError, URLError) as error:
        if isinstance(error, URLError) and not isinstance(
            error.reason,
            ssl.SSLCertVerificationError
        ):
            raise

        context = ssl._create_unverified_context()
        with urlopen(request, timeout=30, context=context) as response:
            data = response.read()

    with open(destination_path, "wb") as file:
        file.write(data)

    return destination_path


def horse_show_office_params(url):
    parsed_url = urlparse(url)
    query = parse_qs(parsed_url.query)
    show_id = query.get("s", [""])[0]
    office_id = query.get("o", [""])[0]

    if not show_id or not office_id:
        raise ValueError("HorseShowOffice URL must include both s= and o= values.")

    return show_id, office_id


def foxvillage_show_id(url):
    parsed_url = urlparse(url)
    query = parse_qs(parsed_url.query)
    show_id = query.get("id", [""])[0]

    if not show_id:
        raise ValueError("FoxVillage URL must include an id= show value.")

    return show_id


def equestrian_hub_show_id(url):
    parsed_url = urlparse(url)
    match = re.search(r"/show/(?P<show_id>\d+)", parsed_url.path)

    if not match:
        raise ValueError("Equestrian Hub URL must look like /show/12345.")

    return match.group("show_id")


def source_type_from_url(url):
    hostname = (urlparse(url).hostname or "").lower()

    if "equestrian-hub.com" in hostname:
        return "equestrianhub"

    if "foxvillage.com" in hostname:
        return "foxvillage"

    if "horseshowoffice.com" in hostname:
        return "horseshowoffice"

    raise ValueError(
        "Unsupported ride-time URL. Use a HorseShowOffice, FoxVillage, "
        "or Equestrian Hub show URL."
    )


def fetch_horse_show_office_rider_links(url):
    show_id, office_id = horse_show_office_params(url)
    html = hso_request(
        urljoin(HORSE_SHOW_OFFICE_BASE_URL, "/hso/ridetimes.asp"),
        {
            "s": show_id,
            "o": office_id,
            "hdnForm": "1",
            "cmdLookup": "View a list of all riders",
        }
    )

    rider_links = {}

    link_pattern = re.compile(
        r"<a\s+href=\"(?P<href>[^\"]*ridetimes\.asp[^\"]*)\"[^>]*>"
        r"(?P<label>.*?)</a>",
        re.IGNORECASE | re.DOTALL
    )

    for match in link_pattern.finditer(html):
        rider = html_cell_text(match.group("label"))

        if not rider or "," not in rider:
            continue

        href = unescape(match.group("href"))
        rider_links[rider] = urljoin(HORSE_SHOW_OFFICE_BASE_URL, href)

    return dict(sorted(rider_links.items(), key=lambda item: item[0].lower()))


def date_to_day_abbreviation(date_text):
    date_text = (date_text or "").strip()

    for date_format in ("%m/%d/%Y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(date_text, date_format).strftime("%a")
        except ValueError:
            pass

    return date_text


def hso_time_to_display_time(time_text):
    time_text = html_cell_text(time_text)
    time_text = time_text.replace(" ", "")

    for time_format in ("%I:%M:%S%p", "%I:%M%p", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(time_text, time_format).strftime("%I:%M %p").lstrip("0")
        except ValueError:
            pass

    return time_text


def parse_horse_show_office_rider_page(html, rider):
    rides = []

    for row_html in re.findall(r"<tr[^>]*>(.*?)</tr>", html, flags=re.IGNORECASE | re.DOTALL):
        cells = [
            html_cell_text(cell)
            for cell in re.findall(r"<td[^>]*>(.*?)</td>", row_html, flags=re.IGNORECASE | re.DOTALL)
        ]
        cells = [cell for cell in cells if cell]

        if len(cells) < 5 or cells[0].lower() == "day":
            continue

        day_text, time_text, class_code, class_name, horse = cells[:5]

        if not re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", day_text):
            continue

        day = date_to_day_abbreviation(day_text)
        time = hso_time_to_display_time(time_text)

        rides.append({
            "rider": rider,
            "day": day,
            "time": time,
            "ready_by": "",
            "coach": "",
            "class": class_code,
            "class_name": class_name,
            "horse": horse,
            "arena": "",
            "arena_number": "",
            "arena_name": "",
            "notes": "",
            "raw": f"{day_text} {time_text} {class_code} {class_name} {horse}",
        })

    return rides


def fetch_horse_show_office_rides(rider_links, riders):
    rides = []

    for rider in riders:
        url = rider_links.get(rider)

        if not url:
            continue

        html = hso_request(url)
        rides.extend(parse_horse_show_office_rider_page(html, rider))

    return rides


def fetch_horse_show_office_class_map(rider_links):
    rides = []

    for rider, url in rider_links.items():
        html = hso_request(url)
        rides.extend(parse_horse_show_office_rider_page(html, rider))

    return class_map_from_rides(rides)


def foxvillage_endpoint(path, **params):
    query = urlencode(params)
    return f"{urljoin(FOXVILLAGE_BASE_URL, path)}?{query}"


def fetch_foxvillage_rider_links(url):
    show_id = foxvillage_show_id(url)
    data = url_request_json(foxvillage_endpoint("/show/GetRiderData", id=show_id))
    rider_links = {}

    for row in data.get("riderData", []):
        rider_name = html_cell_text(row.get("riderName", ""))
        rider_id = str(row.get("riderID", "")).strip()

        if not rider_name or not rider_id:
            continue

        rider_links[rider_name] = {
            "source": "foxvillage",
            "show_id": show_id,
            "rider_id": rider_id,
        }

    return dict(sorted(rider_links.items(), key=lambda item: item[0].lower()))


def fetch_foxvillage_class_map(show_id):
    data = url_request_json(foxvillage_endpoint("/show/GetClassData", id=show_id))
    class_map = {}

    for row in data.get("classData", []):
        class_code = html_cell_text(row.get("classText", ""))
        class_name = html_cell_text(row.get("name", ""))

        if class_code and class_name:
            class_map[class_code] = class_name

    return dict(sorted(class_map.items()))


def parse_foxvillage_rider_rows(rows, rider, class_map):
    rides = []

    for row in rows:
        class_code = html_cell_text(row.get("classText", ""))
        class_name = class_map.get(class_code) or html_cell_text(row.get("test", ""))
        horse = html_cell_text(row.get("horse", ""))
        arena = html_cell_text(row.get("ring", ""))

        if not class_code:
            continue

        rides.append({
            "rider": rider,
            "day": date_to_day_abbreviation(row.get("day", "")),
            "time": hso_time_to_display_time(row.get("rideTime", "")),
            "ready_by": "",
            "coach": "",
            "class": class_code,
            "class_name": class_name,
            "horse": horse,
            "arena": arena,
            "arena_number": "",
            "arena_name": arena,
            "notes": "",
            "raw": " ".join(
                value for value in [
                    html_cell_text(row.get("day", "")),
                    html_cell_text(row.get("rideTime", "")),
                    class_code,
                    class_name,
                    horse,
                    arena,
                ]
                if value
            ),
        })

    return rides


def fetch_foxvillage_rides(rider_links, riders):
    rides = []
    class_maps_by_show = {}

    for rider in riders:
        rider_info = rider_links.get(rider)

        if not rider_info:
            continue

        show_id = rider_info.get("show_id")
        rider_id = rider_info.get("rider_id")

        if not show_id or not rider_id:
            continue

        if show_id not in class_maps_by_show:
            class_maps_by_show[show_id] = fetch_foxvillage_class_map(show_id)

        data = url_request_json(
            foxvillage_endpoint(
                "/show/GetAllRiderData",
                show=show_id,
                id=rider_id
            )
        )
        rows = data.get("riderPageData", [])
        rides.extend(
            parse_foxvillage_rider_rows(
                rows,
                rider,
                class_maps_by_show[show_id]
            )
        )

    return rides


EQUESTRIAN_HUB_MASTERLIST_QUERY = """
query GetShowMasterlist($showId: ID!) {
  masterlist: showMasterlist(showId: $showId) {
    showId
    name
    firstDay
    lastDay
    timeZone
    arenas
    competitions {
      id
      name
      subtitle
      status
      number
      level
      discipline
      arena
      publishingStatus
      numberOfCompetitors
      numberOfFinishedCompetitors
      instant
      assignmentDeadline
    }
    combinations {
      athleteApiId
      horseApiId
      cno
      competitions {
        competitionId
        startTime
        numberInCompetition
        providedMusic
        status
      }
    }
    athletes {
      apiId
      id
      person {
        apiId
        name
      }
    }
    horses {
      apiId
      name
    }
  }
}
"""


def fetch_equestrian_hub_masterlist_by_show_id(show_id):
    if show_id in EQUESTRIAN_HUB_MASTERLIST_CACHE:
        return EQUESTRIAN_HUB_MASTERLIST_CACHE[show_id]

    response = json_request(
        EQUESTRIAN_HUB_GRAPHQL_URL,
        {
            "operationName": "GetShowMasterlist",
            "query": EQUESTRIAN_HUB_MASTERLIST_QUERY,
            "variables": {
                "showId": show_id,
            },
        },
        {
            "Origin": "https://equestrian-hub.com",
            "Referer": f"https://equestrian-hub.com/show/{show_id}",
        }
    )

    if response.get("errors"):
        message = response["errors"][0].get("message", "Unknown GraphQL error")
        raise ValueError(f"Equestrian Hub returned an error: {message}")

    masterlist = response.get("data", {}).get("masterlist")

    if not masterlist:
        raise ValueError("No Equestrian Hub show masterlist was found.")

    EQUESTRIAN_HUB_MASTERLIST_CACHE[show_id] = masterlist
    return masterlist


def fetch_equestrian_hub_masterlist(url):
    return fetch_equestrian_hub_masterlist_by_show_id(
        equestrian_hub_show_id(url)
    )


def fetch_equestrian_hub_rider_links(url):
    show_id = equestrian_hub_show_id(url)
    masterlist = fetch_equestrian_hub_masterlist_by_show_id(show_id)
    rider_links = {}

    for athlete in masterlist.get("athletes", []):
        rider_name = html_cell_text(
            athlete.get("person", {}).get("name", "")
        )
        athlete_api_id = athlete.get("apiId")

        if not rider_name or not athlete_api_id:
            continue

        rider_links[rider_name] = {
            "source": "equestrianhub",
            "show_id": show_id,
            "athlete_api_id": athlete_api_id,
        }

    return dict(sorted(rider_links.items(), key=lambda item: item[0].lower()))


def parse_iso_instant(instant_text, timezone_name=""):
    instant_text = (instant_text or "").strip()

    if not instant_text:
        return None

    try:
        date_time = datetime.fromisoformat(
            instant_text.replace("Z", "+00:00")
        )
    except ValueError:
        return None

    if timezone_name and date_time.tzinfo is not None:
        try:
            date_time = date_time.astimezone(ZoneInfo(timezone_name))
        except Exception:
            pass

    return date_time


def iso_instant_to_day_and_time(instant_text, timezone_name=""):
    date_time = parse_iso_instant(instant_text, timezone_name)

    if not date_time:
        return "", ""

    return (
        date_time.strftime("%a"),
        date_time.strftime("%I:%M %p").lstrip("0"),
    )


def equestrian_hub_class_name(competition):
    class_name = html_cell_text(competition.get("name", ""))
    subtitle = html_cell_text(competition.get("subtitle", ""))

    if subtitle and subtitle not in class_name:
        return f"{class_name} - {subtitle}"

    return class_name


def fetch_equestrian_hub_rides(rider_links, riders):
    if not riders:
        return []

    first_link = next(iter(rider_links.values()), {})
    show_id = first_link.get("show_id")

    if not show_id:
        return []

    masterlist = fetch_equestrian_hub_masterlist_by_show_id(show_id)
    timezone_name = masterlist.get("timeZone", "")
    athletes_by_api_id = {
        athlete.get("apiId"): athlete
        for athlete in masterlist.get("athletes", [])
    }
    horses_by_api_id = {
        horse.get("apiId"): horse
        for horse in masterlist.get("horses", [])
    }
    competitions_by_id = {
        competition.get("id"): competition
        for competition in masterlist.get("competitions", [])
    }
    selected_athlete_ids = {
        rider_links[rider].get("athlete_api_id")
        for rider in riders
        if rider in rider_links
    }
    riders_by_athlete_id = {
        rider_links[rider].get("athlete_api_id"): rider
        for rider in riders
        if rider in rider_links
    }
    rides = []

    for combination in masterlist.get("combinations", []):
        athlete_api_id = combination.get("athleteApiId")

        if athlete_api_id not in selected_athlete_ids:
            continue

        athlete = athletes_by_api_id.get(athlete_api_id, {})
        horse = horses_by_api_id.get(combination.get("horseApiId"), {})
        rider = riders_by_athlete_id.get(athlete_api_id) or html_cell_text(
            athlete.get("person", {}).get("name", "")
        )
        horse_name = html_cell_text(horse.get("name", ""))

        for ride_info in combination.get("competitions", []):
            competition = competitions_by_id.get(
                ride_info.get("competitionId"),
                {}
            )
            class_code = html_cell_text(competition.get("number", ""))

            if not class_code:
                class_code = html_cell_text(ride_info.get("competitionId", ""))

            day, time = iso_instant_to_day_and_time(
                ride_info.get("startTime") or competition.get("instant"),
                timezone_name
            )
            class_name = equestrian_hub_class_name(competition)
            arena = html_cell_text(competition.get("arena", ""))

            rides.append({
                "rider": rider,
                "day": day,
                "time": time,
                "ready_by": "",
                "coach": "",
                "class": class_code,
                "class_name": class_name,
                "horse": horse_name,
                "arena": arena,
                "arena_number": "",
                "arena_name": arena,
                "notes": "",
                "raw": " ".join(
                    value for value in [
                        ride_info.get("startTime", ""),
                        class_code,
                        class_name,
                        horse_name,
                        arena,
                    ]
                    if value
                ),
            })

    return rides


def fetch_rider_links_from_url(url, source_type=None):
    source_type = source_type or source_type_from_url(url)

    if source_type == "equestrianhub":
        return fetch_equestrian_hub_rider_links(url)

    if source_type == "foxvillage":
        return fetch_foxvillage_rider_links(url)

    if source_type == "horseshowoffice":
        return fetch_horse_show_office_rider_links(url)

    raise ValueError("Unsupported show platform.")


def fetch_rides_for_riders(rider_links, riders):
    if not riders:
        return []

    first_link = next(iter(rider_links.values()), "")

    if isinstance(first_link, dict) and first_link.get("source") == "equestrianhub":
        return fetch_equestrian_hub_rides(rider_links, riders)

    if isinstance(first_link, dict) and first_link.get("source") == "foxvillage":
        return fetch_foxvillage_rides(rider_links, riders)

    return fetch_horse_show_office_rides(rider_links, riders)


def class_map_from_rides(rides):
    class_names_by_code = {}

    for ride in rides:
        class_code = ride.get("class", "").strip()
        class_name = ride.get("class_name", "").strip()

        if not class_code or not class_name:
            continue

        names = class_names_by_code.setdefault(class_code, {})
        names[class_name] = names.get(class_name, 0) + 1

    class_map = {}

    for class_code, names in class_names_by_code.items():
        class_map[class_code] = sorted(
            names.items(),
            key=lambda item: (-item[1], len(item[0]), item[0].lower())
        )[0][0]

    return dict(sorted(class_map.items()))


def split_class_and_horse(text, class_code, class_map):
    if text.startswith("Q "):
        text = text[2:].strip()

    class_name = class_map.get(class_code, "")

    if class_name:
        normalized_text = normalize_text(text)
        normalized_class_name = normalize_text(class_name)

        # Try exact class-map match first.
        if normalized_text.startswith(normalized_class_name):
            horse = text[len(class_name):].strip()
            return class_name, horse

        # Try removing common qualifier phrases from the class-map name.
        simplified_class_name = class_name

        qualifier_phrases = [
            "Markel/USEF Qualifying",
            "USEF Qualifying",
            "Qualifying",
        ]

        for phrase in qualifier_phrases:
            simplified_class_name = simplified_class_name.replace(phrase, "").strip()

        simplified_class_name = re.sub(r"\s+", " ", simplified_class_name).strip()
        normalized_simplified = normalize_text(simplified_class_name)

        if normalized_simplified and normalized_text.startswith(normalized_simplified):
            horse = text[len(simplified_class_name):].strip()
            return class_name, horse

    # Fallback patterns for cases where class schedule parsing is imperfect.
    fallback_patterns = [
        r"USEF Developing Horse Grand Prix Test",
        r"USEF Developing Horse Prix St\.? George Test",
        r"USEF 4-Year-Old Test",
        r"USEF Test of Choice(?: (?:Training|First|Second|Third|Fourth) Level)?",
        r"FEI 5-Year-Old Test Preliminary",
        r"FEI 5-Year-Old Test Final",
        r"FEI Pony Team Test",
        r"FEI Pony Individual Test",
        r"FEI Intermediate I",
        r"FEI Intermediate II",
        r"FEI Prix\.? St\. Georges",
        r"Prix St\.? Georges",
        r"FEI Intermediare I",
        r"FEI Intermediare II",
        r"FEI Grand Prix",
        r"FEI Test of Choice",
        r"FEI Freestyle Test of Choice",
        r"USDF Freestyle Test of Choice",
        r"FEI Musical Freestyle Test of Choice",
        r"USDF Musical Freestyle Test of Choice",
        r"Training Level Test \d(?: - (?:Adult Am\.|Open|Jr/Yg Rider))?",
        r"First Level Test \d(?: - (?:Adult Am\.|Open|Jr/Yg Rider))?",
        r"Second Level Test \d(?: - (?:Adult Am\.|Open|Jr/Yg Rider))?",
        r"Third Level Test \d(?: - (?:Adult Am\.|Open|Jr/Yg Rider))?",
        r"Fourth Level Test \d(?: - (?:Adult Am\.|Open|Jr/Yg Rider))?",
        r"USDF Introductory Test [ABC]",
        r"USEF Pony Test of Choice",
        r"Dressage Seat Equitation - Adult Am\.",
        r"Dressage Seat Equitation - U16",
        r"Dressage Seat Equitation - Open",
        r"Dressage Seat Equitation - 17-21",
        r"Dressage Seat Equitation",
        r"IBC - .*",
        r"Stars and Stripes Benefit TOC \(USEF\)",
        r"Stars and Stripes Benefit TOC \(FEI\)",
        r"Stars and Stripes Benefit TOC \(Intro\)",
        r"Materiale 3/4/5 Year old Fillies/Mares",
        r"Materiale 3/4/5 year old Colts/Geldings/Stallions",
        r"Materiale 4- & 5-Year Old Fillies",
        r"Materiale 4- & 5-Year Old Stallions/Geldings",
        r"Young Horse Test of Choice \(not including 7 year olds\)",
        r"Young Horse Test of Choice \(excl 7YO\)",
    ]

    for pattern in fallback_patterns:
        match = re.match(pattern, text)
        if match:
            found_class_name = match.group(0).strip()
            horse = text[match.end():].strip()
            return class_name or found_class_name, horse

    return class_name, text


def ride_time_line_is_continuation(line):
    line = line.strip()

    if not line:
        return False

    if ride_pattern.match(line) or rider_pattern.match(line):
        return False

    if line.startswith("Day Ride Time"):
        return False

    if line.startswith("* = "):
        return False

    ignored_lines = {
        "Ride Times by Rider",
        "USEF Number: USDF Number:",
    }

    return line not in ignored_lines


def combine_ride_time_continuation_lines(lines):
    combined_lines = []

    for line in lines:
        stripped_line = line.strip()

        if (
            combined_lines
            and ride_time_line_is_continuation(stripped_line)
            and ride_pattern.match(combined_lines[-1])
        ):
            previous_line = combined_lines[-1]
            arena_match = arena_pattern.search(previous_line)

            if not arena_match:
                arena_match = named_arena_pattern.search(previous_line)

            if arena_match:
                combined_lines[-1] = (
                    f"{previous_line[:arena_match.start()].strip()} "
                    f"{stripped_line} "
                    f"{previous_line[arena_match.start():].strip()}"
                )
            else:
                combined_lines[-1] = f"{previous_line} {stripped_line}"
        else:
            combined_lines.append(line)

    return combined_lines


def extract_horse_and_arena(line, class_code, class_map):
    arena_match = arena_pattern.search(line)

    if arena_match:
        before_arena = line[:arena_match.start()].strip()
        arena = line[arena_match.start():].strip()
    else:
        arena_match = named_arena_pattern.search(line)

        if not arena_match:
            return "", "", ""

        before_arena = line[:arena_match.start()].strip()
        arena = arena_match.group("arena").strip()

    parts = before_arena.split()

    if len(parts) < 5:
        return "", "", arena

    remaining = " ".join(parts[4:]).strip()

    class_name, horse = split_class_and_horse(remaining, class_code, class_map)

    return class_name, horse, arena


def split_arena(arena):
    match = re.match(r"(?P<number>\d+):\s*(?P<name>.+)", arena)

    if match:
        return match.group("number"), match.group("name")

    if arena.strip() == "0":
        return "0", ""

    named_arena = re.sub(r"\s*\([^)]*\)\s*$", "", arena).strip()

    return "", named_arena


def parse_rides(lines, my_riders, class_map):
    rides = []
    current_rider = None

    lines = combine_ride_time_continuation_lines(lines)

    for line in lines:
        line = line.strip()

        if rider_pattern.match(line):
            current_rider = line
            continue

        if current_rider not in my_riders:
            continue

        if ride_pattern.match(line):
            parts = line.split()

            if len(parts) < 4:
                continue

            day = parts[0]
            time = parts[1] + " " + parts[2].replace("*", "")
            class_code = parts[3]

            class_name, horse, arena = extract_horse_and_arena(
                line,
                class_code,
                class_map
            )

            arena_number, arena_name = split_arena(arena)

            rides.append({
                "rider": current_rider,
                "day": day,
                "time": time,
                "ready_by": "",
                "coach": "",
                "class": class_code,
                "class_name": class_name,
                "horse": horse,
                "arena": arena,
                "arena_number": arena_number,
                "arena_name": arena_name,
                "notes": "",
                "raw": line
            })

    return rides


def print_validation_report(rides, class_map):
    missing_class_codes = sorted({
        r["class"]
        for r in rides
        if not r["class_name"]
    })

    print("\n===== VALIDATION REPORT =====")

    print(f"Class definitions loaded: {len(class_map)}")

    if missing_class_codes:
        print("\nMissing class definitions:")
        for code in missing_class_codes:
            print(f"- {code}")
    else:
        print("\nNo missing class definitions found.")

    rides_missing_horse = [
        r for r in rides
        if not r["horse"]
    ]

    if rides_missing_horse:
        print("\nRides missing horse name:")
        for r in rides_missing_horse:
            print(f"- {r['day']} {r['time']} {r['rider']} {r['class']}")
    else:
        print("No rides missing horse names.")

    rides_missing_arena = [
        r for r in rides
        if not r["arena_number"] and not r["arena_name"]
    ]

    if rides_missing_arena:
        print("\nRides missing arena:")
        for r in rides_missing_arena:
            print(f"- {r['day']} {r['time']} {r['rider']} {r['class']}")
    else:
        print("No rides missing arena info.")

    print("=============================\n")


def make_ride_id(ride):
    raw_id = (
        f"{ride['day']}_"
        f"{ride['time']}_"
        f"{ride['rider']}_"
        f"{ride['horse']}_"
        f"{ride['class']}"
    )

    # Keep only letters and numbers so AppSheet has a clean key
    clean_id = re.sub(r"[^A-Za-z0-9]+", "_", raw_id)
    clean_id = clean_id.strip("_")

    return clean_id


def export_missing_class_definitions(rides):
    output_file = supporting_output_path("missing_class_definitions.csv")

    missing = []

    for r in rides:
        if not r["class_name"]:
            missing.append(r)

    with open(output_file, "w", newline="") as file:
        writer = csv.writer(file)

        writer.writerow([
            "Class #",
            "Day",
            "Ride Time",
            "Rider",
            "Horse",
            "Raw Line"
        ])

        for r in missing:
            writer.writerow([
                r["class"],
                r["day"],
                r["time"],
                r["rider"],
                r["horse"],
                r["raw"]
            ])

    if missing:
        print(f"Missing class definitions exported to: {output_file}")
    else:
        print("No missing class definitions file needed.")


def export_used_class_codes(rides, class_map):
    output_file = supporting_output_path("used_class_codes.csv")

    used_codes = sorted({
        r["class"]
        for r in rides
    })

    with open(output_file, "w", newline="") as file:
        writer = csv.writer(file)

        writer.writerow([
            "Class #",
            "Found In Class Map?",
            "Class Name"
        ])

        for code in used_codes:
            writer.writerow([
                code,
                "Yes" if code in class_map else "No",
                class_map.get(code, "")
            ])

    print(f"Used class codes exported to: {output_file}")


def ride_sort_key(ride):
    ride_time = datetime.strptime(ride["time"], "%I:%M %p").time()
    return (DAY_ORDER.get(ride["day"], 99), ride_time)


def export_schedule_csv(rides):
    output_file = supporting_output_path("barn_schedule.csv")

    with open(output_file, "w", newline="") as file:
        writer = csv.writer(file)

        writer.writerow([
            "Day",
            "Ride Time",
            "Ready By / On Horse By",
            "Coach",
            "Rider",
            "Horse",
            "Class #",
            "Class Name",
            "Arena #",
            "Arena Name",
            "Notes"
        ])

        for r in rides:
            writer.writerow([
                r["day"],
                r["time"],
                r["ready_by"],
                r["coach"],
                r["rider"],
                r["horse"],
                r["class"],
                r["class_name"],
                r["arena_number"],
                r["arena_name"],
                r["notes"]
            ])

    print(f"\nSchedule exported to: {output_file}")

def setup_schedule_sheet(sheet, rides, title):
    headers = [
        "Day",
        "Ride Time",
        "Ready By / On Horse By",
        "Coach",
        "Rider",
        "Horse",
        "Class #",
        "Class Name",
        "Arena #",
        "Arena Name",
        "Notes"
    ]

    # Title row
    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=16)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")

    # Merge title across the full table width
    sheet.merge_cells("A1:K1")

    # Blank spacer row
    sheet.append([])

    # Header row starts on row 3
    sheet.append(headers)

    for r in rides:
        sheet.append([
            r["day"],
            r["time"],
            r["ready_by"],
            r["coach"],
            r["rider"],
            r["horse"],
            r["class"],
            r["class_name"],
            r["arena_number"],
            r["arena_name"],
            r["notes"]
        ])

    # Header formatting
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for cell in sheet[3]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Freeze title + spacer + header
    sheet.freeze_panes = "A4"

    # Auto-filter on table only
    sheet.auto_filter.ref = f"A3:K{sheet.max_row}"

    # Column widths
    column_widths = {
        "A": 10,
        "B": 12,
        "C": 22,
        "D": 24,
        "E": 24,
        "F": 24,
        "G": 12,
        "H": 38,
        "I": 10,
        "J": 34,
        "K": 30,
        "L": 20,
    }

    for column_letter, width in column_widths.items():
        sheet.column_dimensions[column_letter].width = width

    # Body formatting
    for row in sheet.iter_rows(min_row=4):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Highlight manual entry columns
    for row in range(4, sheet.max_row + 1):
        sheet[f"C{row}"].fill = PatternFill("solid", fgColor="FFF2CC")  #Ready By / On Horse By
        sheet[f"D{row}"].fill = PatternFill("solid", fgColor="FFF2CC")  #Coach
        sheet[f"K{row}"].fill = PatternFill("solid", fgColor="FFF2CC")  #Notes

    # Slightly taller title row
    sheet.row_dimensions[1].height = 24

    # Page setup for printing
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    # Repeat title/header area when printed
    sheet.print_title_rows = "1:3"


def export_appsheet_csv(rides):
    output_file = supporting_output_path("appsheet_schedule.csv")

    with open(output_file, "w", newline="") as file:
        writer = csv.writer(file)

        writer.writerow([
            "Ride ID",
            "Show Name",
            "Day",
            "Ride Time",
            "Ready By / On Horse By",
            "Coach",
            "Rider",
            "Horse",
            "Class #",
            "Class Name",
            "Arena #",
            "Arena Name",
            "Notes",
            "Status"
        ])

        for r in rides:
            writer.writerow([
                make_ride_id(r),
                "",                 # Show Name - fill in later if desired
                r["day"],
                r["time"],
                "",                 # Ready By / On Horse By - editable in AppSheet
                r["coach"],
                r["rider"],
                r["horse"],
                r["class"],
                r["class_name"],
                r["arena_number"],
                r["arena_name"],
                "",                 # Notes - editable in AppSheet
                "Scheduled"         # Status - editable in AppSheet
            ])

    print(f"AppSheet schedule exported to: {output_file}")


def export_schedule_xlsx(rides):
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    output_file = output_path("barn_schedule.xlsx")

    workbook = Workbook()

    # Main all-rides sheet
    all_sheet = workbook.active
    all_sheet.title = "All Rides"
    setup_schedule_sheet(all_sheet, rides, f"{SHOW_NAME} - All Rides Barn Schedule")

    # Separate sheets by day
    day_titles = {
        "Mon": "Monday Barn Schedule",
        "Tue": "Tuesday Barn Schedule",
        "Wed": "Wednesday Barn Schedule",
        "Thu": "Thursday Barn Schedule",
        "Fri": "Friday Barn Schedule",
        "Sat": "Saturday Barn Schedule",
        "Sun": "Sunday Barn Schedule",
    }

    days = []
    for r in rides:
        if r["day"] not in days:
            days.append(r["day"])

    for day in days:
        day_rides = [
            r for r in rides
            if r["day"] == day
        ]

        day_sheet = workbook.create_sheet(title=day)
        setup_schedule_sheet(
            day_sheet,
            day_rides,
            f"{SHOW_NAME} - {day_titles.get(day, f'{day} Barn Schedule')}"
        )

    workbook.save(output_file)

    print(f"Formatted Excel schedule exported to: {output_file}")


def export_rides(rides, class_map):
    rides.sort(key=ride_sort_key)

    print_validation_report(rides, class_map)
    export_missing_class_definitions(rides)
    export_used_class_codes(rides, class_map)

    print(f"\nFiltered rides: {len(rides)}\n")

    current_day = None

    for r in rides:
        if r["day"] != current_day:
            current_day = r["day"]
            print(f"\n===== {current_day.upper()} =====")

        print(
            f"{r['time']} — "
            f"{r['rider']} — "
            f"{r['horse']} — "
            f"{r['class']} — "
            f"{r['class_name']} — "
            f"Arena {r['arena_number']}: {r['arena_name']}"
        )

    export_schedule_csv(rides)
    export_schedule_xlsx(rides)
    export_appsheet_csv(rides)


def main(class_map_override=None):
    my_riders = load_riders()

    if not my_riders:
        print("No riders found. Add riders to riders.txt")
        return

    if class_map_override is not None:
        class_map = class_map_override
    else:
        class_map = build_class_map()

    print(f"Loaded {len(class_map)} class definitions from class schedule PDFs.")
    export_class_map_csv(class_map)

    lines = extract_lines_from_folder(PDF_FOLDER)
    rides = parse_rides(lines, my_riders, class_map)
    export_rides(rides, class_map)
    
if __name__ == "__main__":
    main()
